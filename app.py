import streamlit as st
import pandas as pd
import openpyxl
import re, io
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict

# ══════════════════════════════════════════════════════════════════════════════
# PAGE CONFIG & STYLES
# ══════════════════════════════════════════════════════════════════════════════
st.set_page_config(page_title="Rock Point Camp Group Sorter", page_icon="⛪", layout="wide")
st.markdown("""
<style>
/* ── Google Fonts ────────────────────────────────────────────────────────────── */
@import url('https://fonts.googleapis.com/css2?family=Montserrat:wght@400;600;700;900&family=Inter:wght@300;400;500;700;900&display=swap');

/* ── Design tokens — KIDS (default) ─────────────────────────────────────────── */
:root {
  --primary:     #152B50;
  --accent1:     #52B8DE;
  --accent2:     #E25D32;
  --accent3:     #F3B73F;
  --accent4:     #8AB738;
  --bg:          #F4F6FB;
  --surface:     #FFFFFF;
  --border:      #DDE3F0;
  --text:        #152B50;
  --muted:       #6B7A99;
  --radius:      12px;
  --font-h:      'Montserrat', sans-serif;
  --font-b:      'Montserrat', sans-serif;
  --btn:         #152B50;
  --btn-hov:     #1E3E72;
  --accent-step: #52B8DE;
}

/* ── Base ───────────────────────────────────────────────────────────────────── */
html, body, [class*="css"] { font-family: var(--font-b) !important; }
.block-container {
  max-width: 880px !important;
  padding: 0 1.5rem 3rem !important;
  margin: 0 auto;
}
section[data-testid="stSidebar"] { display: none; }
#MainMenu, footer, header { visibility: hidden; }

/* ── App shell: two-column layout ──────────────────────────────────────────── */
.rp-shell {
  display: grid;
  grid-template-columns: 260px 1fr;
  gap: 0;
  min-height: 100vh;
  background: var(--bg);
}

/* ── Left sidebar panel ─────────────────────────────────────────────────────── */
.rp-sidebar {
  background: var(--primary);
  padding: 2rem 1.5rem;
  display: flex;
  flex-direction: column;
  gap: 0;
  position: relative;
  overflow: hidden;
}
/* subtle diagonal texture */
.rp-sidebar::before {
  content: '';
  position: absolute;
  inset: 0;
  background: repeating-linear-gradient(
    135deg,
    transparent,
    transparent 28px,
    rgba(255,255,255,.025) 28px,
    rgba(255,255,255,.025) 30px
  );
  pointer-events: none;
}
/* accent bar bottom of sidebar */
.rp-sidebar::after {
  content: '';
  position: absolute;
  bottom: 0; left: 0; right: 0;
  height: 4px;
}
/* KIDS: rainbow bar */
.kids-sidebar::after {
  background: linear-gradient(to right, #E25D32, #F3B73F, #52B8DE, #8AB738);
}
/* YTH: cyan bar */
.yth-sidebar::after { background: #1EBFE1; }

.rp-logo-wrap {
  margin-bottom: 2rem;
  display: flex;
  flex-direction: column;
  align-items: flex-start;
  gap: .5rem;
}
.rp-church-label {
  font-family: var(--font-h);
  font-size: .62rem;
  font-weight: 700;
  letter-spacing: 3px;
  text-transform: uppercase;
  color: rgba(255,255,255,.45);
}
/* Logo SVG container */
.rp-logo-svg {
  width: 180px;
  height: auto;
}
.rp-logo-svg svg { width: 100%; height: auto; display: block; }

.rp-tagline {
  font-size: .78rem;
  color: rgba(255,255,255,.5);
  line-height: 1.5;
  margin-top: auto;
  padding-top: 2rem;
  font-style: italic;
}

/* ── Right main panel ───────────────────────────────────────────────────────── */
.rp-main {
  padding: 2rem 2rem 3rem;
  background: var(--bg);
}

/* ── Camp type switcher ─────────────────────────────────────────────────────── */
.camp-switcher {
  display: flex;
  gap: 8px;
  margin-bottom: 1.8rem;
  border-bottom: 2px solid var(--border);
  padding-bottom: 1.2rem;
}
.camp-tab {
  flex: 1;
  padding: .75rem 1rem;
  border-radius: var(--radius);
  border: 2px solid var(--border);
  background: var(--surface);
  cursor: pointer;
  text-align: center;
  font-family: var(--font-h);
  font-weight: 700;
  font-size: .88rem;
  letter-spacing: .3px;
  color: var(--muted);
  position: relative;
  overflow: hidden;
  transition: all .15s;
}
.camp-tab::before {
  content: '';
  position: absolute;
  top: 0; left: 0; right: 0;
  height: 3px;
  opacity: 0;
  transition: opacity .15s;
}
.camp-tab.kids::before { background: linear-gradient(to right,#E25D32,#F3B73F,#52B8DE,#8AB738); }
.camp-tab.yth::before  { background: #1EBFE1; }
.camp-tab.active { border-color: var(--primary); color: var(--primary); }
.camp-tab.kids.active::before, .camp-tab.yth.active::before { opacity: 1; }

/* ── Section headers ────────────────────────────────────────────────────────── */
.sec-head {
  display: flex;
  align-items: center;
  gap: .7rem;
  margin: 1.6rem 0 .8rem;
}
.sec-num {
  width: 26px; height: 26px;
  border-radius: 50%;
  background: var(--accent-step);
  color: var(--primary);
  font-family: var(--font-h);
  font-size: .75rem;
  font-weight: 900;
  display: flex; align-items: center; justify-content: center;
  flex-shrink: 0;
}
.yth-sec .sec-num { border-radius: 4px; background: #1EBFE1; color: #1E1E1E; }
.sec-title {
  font-family: var(--font-h);
  font-size: .72rem;
  font-weight: 700;
  text-transform: uppercase;
  letter-spacing: 1.6px;
  color: var(--muted);
}

/* ── Upload cards ───────────────────────────────────────────────────────────── */
.upload-grid {
  display: grid;
  grid-template-columns: 1fr 1fr;
  gap: 12px;
  margin-bottom: 1rem;
}
.upload-grid.single { grid-template-columns: 1fr; }
.upload-card {
  background: var(--surface);
  border: 1.5px solid var(--border);
  border-radius: var(--radius);
  padding: 1.1rem 1.3rem;
  position: relative;
}
.upload-card::before {
  content: '';
  position: absolute;
  top: 0; left: 0;
  width: 4px; height: 100%;
  border-radius: var(--radius) 0 0 var(--radius);
}
.upload-card.reg::before  { background: var(--accent1); }
.upload-card.cg::before   { background: var(--accent2); }
.upload-card.ldr::before  { background: var(--accent4); }
.upload-card.supp::before { background: var(--accent3); }
.upload-card.opt  { border-style: dashed; opacity: .85; }
.upload-label {
  font-family: var(--font-h);
  font-size: .68rem;
  font-weight: 700;
  text-transform: uppercase;
  letter-spacing: 1.3px;
  color: var(--muted);
  margin-bottom: .5rem;
  display: flex;
  align-items: center;
  gap: .4rem;
}
.upload-label .badge {
  background: var(--accent1);
  color: var(--primary);
  border-radius: 4px;
  padding: 1px 6px;
  font-size: .6rem;
  font-weight: 900;
  letter-spacing: .5px;
}
.upload-label .badge.opt {
  background: var(--border);
  color: var(--muted);
}
.upload-desc {
  font-size: .76rem;
  color: var(--muted);
  margin-bottom: .7rem;
  line-height: 1.4;
}

/* ── Param cards ────────────────────────────────────────────────────────────── */
.param-card {
  background: var(--surface);
  border: 1.5px solid var(--border);
  border-radius: var(--radius);
  padding: 1.2rem 1.4rem;
  margin-bottom: .9rem;
}
.kids-param { border-left: 4px solid var(--accent1); }
.yth-param  { border-top:  3px solid #1EBFE1; }
.param-card h4 {
  font-family: var(--font-h);
  font-size: .68rem;
  font-weight: 700;
  text-transform: uppercase;
  letter-spacing: 1.3px;
  color: var(--muted);
  margin: 0 0 .9rem;
  padding-bottom: .6rem;
  border-bottom: 1px solid var(--border);
}

/* ── Toggle row ─────────────────────────────────────────────────────────────── */
.toggle-row { display: flex; gap: 1.5rem; flex-wrap: wrap; }

/* ── Banners ────────────────────────────────────────────────────────────────── */
.ok-banner   { background:#F0FDF4; border-left:4px solid #8AB738; border-radius:0 8px 8px 0;
               padding:.7rem 1rem; color:#14532d; font-size:.85rem; margin:.5rem 0; }
.warn-banner { background:#FFFBEB; border-left:4px solid #F3B73F; border-radius:0 8px 8px 0;
               padding:.7rem 1rem; color:#7d5a00; font-size:.85rem; margin:.5rem 0; }
.info-banner { background:#EFF8FF; border-left:4px solid #52B8DE; border-radius:0 8px 8px 0;
               padding:.7rem 1rem; color:#0c5c82; font-size:.85rem; margin:.5rem 0; }

/* ── Stats ──────────────────────────────────────────────────────────────────── */
.stats-row { display:grid; grid-template-columns:repeat(4,1fr); gap:10px; margin:1.4rem 0; }
.stat-box {
  background: var(--surface);
  border: 1.5px solid var(--border);
  border-radius: var(--radius);
  padding: 14px 8px; text-align: center;
  position: relative; overflow: hidden;
}
.stat-box::after { content:''; position:absolute; bottom:0; left:0; right:0;
                   height:3px; background:var(--accent1); opacity:.35; }
.stat-box .val { font-family:var(--font-h); font-size:1.8rem; font-weight:900;
                 color:var(--primary); line-height:1; }
.stat-box .lbl { font-size:.63rem; color:var(--muted); margin-top:5px;
                 text-transform:uppercase; letter-spacing:.8px; font-weight:700; }
.stat-box.warn { border-color:#F3B73F; }
.stat-box.warn .val { color:#b8600a; }
.stat-box.warn::after { background:#F3B73F; opacity:1; }
.stat-box.good { border-color:#8AB738; }
.stat-box.good .val { color:#2e7d32; }
.stat-box.good::after { background:#8AB738; opacity:1; }

/* ── Results table ──────────────────────────────────────────────────────────── */
.grade-table { width:100%; border-collapse:collapse; font-size:.85rem; margin-top:.6rem; }
.grade-table th { background:var(--primary); color:#fff; padding:9px 14px; text-align:left;
                  font-family:var(--font-h); font-weight:700; font-size:.68rem;
                  letter-spacing:1px; text-transform:uppercase; }
.grade-table td { padding:9px 14px; border-bottom:1px solid var(--border); }
.grade-table tr:last-child td { border-bottom:none; }
.grade-table tr:hover td { background:#f8faff; }
.grade-dot { display:inline-block; width:8px; height:8px; border-radius:50%; margin-right:8px; }

/* ── Download & Generate buttons ────────────────────────────────────────────── */
.stDownloadButton > button {
  width:100% !important;
  background: var(--btn) !important;
  color:#fff !important; border:none !important;
  border-radius:8px !important;
  padding:.8rem 1rem !important;
  font-family:var(--font-h) !important;
  font-size:.9rem !important; font-weight:700 !important;
  letter-spacing:.5px !important; margin-top:.5rem;
}
.stDownloadButton > button:hover { background: var(--btn-hov) !important; }
.stButton > button[kind="primary"] {
  background: var(--btn) !important; color:#fff !important;
  border:none !important; border-radius:8px !important;
  font-family:var(--font-h) !important; font-weight:700 !important;
  letter-spacing:.5px !important;
}
.stButton > button[kind="primary"]:hover { background: var(--btn-hov) !important; }

/* ── File uploader style ────────────────────────────────────────────────────── */
[data-testid="stFileUploader"] {
  border: 1.5px dashed var(--border) !important;
  border-radius:10px !important;
}

/* ── Misc ───────────────────────────────────────────────────────────────────── */
hr { border-color: var(--border) !important; margin:1.2rem 0 !important; }
div[data-testid="stRadio"] label { font-size:.9rem !important; }
div[data-testid="stToggle"] label { font-size:.9rem !important; }
[data-testid="stSlider"] [role="slider"] { background:var(--accent1) !important; }
.stCaption { color:var(--muted) !important; font-size:.78rem !important; }
</style>
""", unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════════════════════
# SVG LOGOS  (inline SVG — no external files needed)
# ══════════════════════════════════════════════════════════════════════════════
# KIDS logo: "ROCK POINT" wordmark above bold "KIDS" with the diagonal K slash
KIDS_LOGO_SVG = """
<svg viewBox="0 0 220 80" xmlns="http://www.w3.org/2000/svg" class="rp-logo-svg">
  <!-- ROCK POINT small label -->
  <text x="2" y="14" font-family="Montserrat,sans-serif" font-weight="900"
        font-size="9" letter-spacing="3.5" fill="rgba(255,255,255,0.55)"
        text-anchor="start">ROCK POINT</text>
  <!-- K -->
  <text x="2" y="68" font-family="Montserrat,sans-serif" font-weight="900"
        font-size="52" fill="#E25D32">K</text>
  <!-- diagonal slash on K -->
  <line x1="18" y1="22" x2="38" y2="68" stroke="#F3B73F" stroke-width="5"
        stroke-linecap="round"/>
  <!-- I -->
  <text x="44" y="68" font-family="Montserrat,sans-serif" font-weight="900"
        font-size="52" fill="#F3B73F">I</text>
  <!-- D -->
  <text x="62" y="68" font-family="Montserrat,sans-serif" font-weight="900"
        font-size="52" fill="#52B8DE">D</text>
  <!-- S -->
  <text x="104" y="68" font-family="Montserrat,sans-serif" font-weight="900"
        font-size="52" fill="#8AB738">S</text>
</svg>"""

# YTH logo: large bold "YTH" with cross circle icon
YTH_LOGO_SVG = """
<svg viewBox="0 0 200 80" xmlns="http://www.w3.org/2000/svg" class="rp-logo-svg">
  <!-- Cross circle -->
  <circle cx="16" cy="18" r="14" fill="#1EBFE1"/>
  <line x1="16" y1="8"  x2="16" y2="28" stroke="#1E1E1E" stroke-width="3.5"
        stroke-linecap="round"/>
  <line x1="10" y1="14" x2="22" y2="14" stroke="#1E1E1E" stroke-width="3.5"
        stroke-linecap="round"/>
  <!-- YTH text -->
  <text x="36" y="30" font-family="Inter,Helvetica Neue,sans-serif" font-weight="900"
        font-size="28" letter-spacing="-1" fill="#FFFFFF">YTH</text>
  <!-- ROCK POINT below -->
  <text x="36" y="46" font-family="Inter,Helvetica Neue,sans-serif" font-weight="400"
        font-size="9" letter-spacing="3" fill="rgba(255,255,255,0.45)">ROCK POINT</text>
</svg>"""

# ══════════════════════════════════════════════════════════════════════════════
# CONSTANTS  (shared by UI, sorting logic, and Excel builder)
# ══════════════════════════════════════════════════════════════════════════════
WHITE='FFFFFF'; BLACK='000000'; HELPER_BG='EBF3FB'; HELPER_HDR='1F618D'

FRIEND_REQUEST_VARIANTS = [
    'Friend Request', 'Friend Requests',
    "Youth's Friend Request", "Youth's Friend Requests",
    "Youth\u2019s Friend Request", "Youth\u2019s Friend Requests",
]

KEEP_COLS_PREF = ['First Name','Last Name','Grade','Gender','Community Group',
                  'T-Shirt Size','Friend Request','Mobile Phone Number',
                  'Home Phone Number','Home Email Address',
                  'List Medical Conditions','List Behavioral Concerns',
                  'Emergency Contact','Registration Contact First Name',
                  'Registration Contact Last Name','Registration Contact Phone Number']

PALETTE = [
    '1A5276','1E8449','784212','6C3483','117A65','922B21','2471A3','C0392B',
    '0E6655','7D6608','4A235A','1B4F72','145A32','7B241C','1A5276','6E2FBF',
]
def bucket_color(idx): return PALETTE[idx % len(PALETTE)]
def parse_bucket(key):
    parts = key.rsplit('_', 1)
    return (parts[0], parts[1]) if len(parts) == 2 else (key, '')

# ══════════════════════════════════════════════════════════════════════════════
# STEP 1 — CAMP TYPE  (runs first so header can theme correctly)
# ══════════════════════════════════════════════════════════════════════════════
camp_type = st.radio(
    "Camp type",
    options=['KIDS', 'YTH'],
    horizontal=True,
    label_visibility='collapsed',
    key='camp_type_sel',
)
if 'camp_type_prev' not in st.session_state:
    st.session_state['camp_type_prev'] = camp_type
elif st.session_state['camp_type_prev'] != camp_type:
    st.session_state['camp_type_prev'] = camp_type
    st.rerun()

_is_yth   = (camp_type == 'YTH')
_svar     = 'yth' if _is_yth else 'kids'
_font     = 'Inter,sans-serif' if _is_yth else 'Montserrat,sans-serif'
_sec_cls  = 'yth-sec' if _is_yth else ''
_pcard    = 'yth-param' if _is_yth else 'kids-param'
_logo_svg = YTH_LOGO_SVG if _is_yth else KIDS_LOGO_SVG
_sidebar_cls = 'yth-sidebar' if _is_yth else 'kids-sidebar'
_tagline  = ("Creating purposeful environments\nthat are community driven"
             if _is_yth else
             "Helping kids learn about a God\nthey can't see through relationships\nwith people they can see")

# Camp type radio is hidden (we render our own visual switcher below)
# but we still need its value — just hide the Streamlit widget
st.markdown('<style>div[data-testid="stRadio"]{display:none!important;}</style>',
            unsafe_allow_html=True)

# ── Sidebar HTML ──────────────────────────────────────────────────────────────
st.markdown(f"""
<div class="rp-sidebar {_sidebar_cls}" style="position:fixed;top:0;left:0;width:260px;
     height:100vh;z-index:10;display:flex;flex-direction:column;padding:2rem 1.6rem;
     background:#152B50;overflow:hidden;">
  <div style="position:absolute;inset:0;background:repeating-linear-gradient(135deg,
    transparent,transparent 28px,rgba(255,255,255,.025) 28px,
    rgba(255,255,255,.025) 30px);pointer-events:none;"></div>
  {'<div style="position:absolute;bottom:0;left:0;right:0;height:4px;background:linear-gradient(to right,#E25D32,#F3B73F,#52B8DE,#8AB738);"></div>' if not _is_yth else
   '<div style="position:absolute;bottom:0;left:0;right:0;height:4px;background:#1EBFE1;"></div>'}
  <div style="position:relative;z-index:1;">
    {_logo_svg}
    <div style="margin-top:2.5rem;">
      <div style="font-family:{_font};font-size:.6rem;font-weight:700;letter-spacing:2.5px;
           text-transform:uppercase;color:rgba(255,255,255,.35);margin-bottom:.8rem;">
        GROUP SORTER
      </div>
      <div style="font-family:{_font};font-size:.75rem;color:rgba(255,255,255,.45);
           line-height:1.6;font-style:italic;">
        {_tagline.replace(chr(10),'<br>')}
      </div>
    </div>
    <div style="margin-top:2.5rem;border-top:1px solid rgba(255,255,255,.1);padding-top:1.5rem;">
      <div style="font-family:{_font};font-size:.6rem;font-weight:700;letter-spacing:2px;
           text-transform:uppercase;color:rgba(255,255,255,.3);margin-bottom:.8rem;">
        Camp Type
      </div>
      <div style="display:flex;flex-direction:column;gap:6px;">
        <div style="padding:.5rem .8rem;border-radius:8px;font-family:{_font};font-size:.8rem;
             font-weight:700;{'background:rgba(82,184,222,.2);color:#52B8DE;border:1.5px solid rgba(82,184,222,.3);' if not _is_yth else 'color:rgba(255,255,255,.3);border:1.5px solid rgba(255,255,255,.08);'}">
          🏕️ &nbsp;KIDS Camp
        </div>
        <div style="padding:.5rem .8rem;border-radius:8px;font-family:{_font};font-size:.8rem;
             font-weight:700;{'background:rgba(30,191,225,.2);color:#1EBFE1;border:1.5px solid rgba(30,191,225,.3);' if _is_yth else 'color:rgba(255,255,255,.3);border:1.5px solid rgba(255,255,255,.08);'}">
          ✝️ &nbsp;YTH Camp
        </div>
      </div>
    </div>
  </div>
</div>
<div style="margin-left:260px;">
""", unsafe_allow_html=True)

# ── Main content area (offset for sidebar) ────────────────────────────────────
st.markdown(f"""
<div style="padding:1.8rem 1rem .5rem;">
  <div style="display:flex;align-items:center;gap:1rem;margin-bottom:.4rem;">
    <div style="font-family:{_font};font-size:1.5rem;font-weight:900;
         color:#152B50;letter-spacing:-.5px;">
      {'YTH Camp' if _is_yth else 'KIDS Camp'} — Group Sorter
    </div>
    <div style="padding:3px 10px;border-radius:20px;font-size:.7rem;font-weight:700;
         font-family:{_font};letter-spacing:.5px;
         {'background:#1EBFE1;color:#1E1E1E;' if _is_yth else 'background:#E25D32;color:#fff;'}">
      {'YTH' if _is_yth else 'KIDS'}
    </div>
  </div>
  <div style="font-size:.82rem;color:#6B7A99;font-family:{_font};">
    Upload your files, configure parameters, and generate sorted camp groups.
  </div>
</div>
""", unsafe_allow_html=True)

# ── Camp type switcher pills ──────────────────────────────────────────────────
kids_active = 'border:2px solid #152B50;background:#fff;color:#152B50;' if not _is_yth else 'border:2px solid #DDE3F0;background:#F4F6FB;color:#999;'
yth_active  = 'border:2px solid #1E1E1E;background:#1E1E1E;color:#fff;' if _is_yth  else 'border:2px solid #DDE3F0;background:#F4F6FB;color:#999;'
st.markdown(f"""
<div style="display:flex;gap:8px;margin:0 1rem 1.4rem;padding-bottom:1.2rem;
     border-bottom:1.5px solid #DDE3F0;">
  <button onclick="window.location.href='?camp_type_sel=KIDS'"
    style="flex:1;padding:.65rem;border-radius:10px;cursor:pointer;font-family:{_font};
    font-weight:700;font-size:.85rem;{kids_active}position:relative;overflow:hidden;">
    {'<div style="position:absolute;top:0;left:0;right:0;height:3px;background:linear-gradient(to right,#E25D32,#F3B73F,#52B8DE,#8AB738);"></div>' if not _is_yth else ''}
    🏕️ &nbsp;KIDS Camp
  </button>
  <button onclick="window.location.href='?camp_type_sel=YTH'"
    style="flex:1;padding:.65rem;border-radius:10px;cursor:pointer;font-family:{_font};
    font-weight:700;font-size:.85rem;{yth_active}position:relative;overflow:hidden;">
    {'<div style="position:absolute;top:0;left:0;right:0;height:3px;background:#1EBFE1;"></div>' if _is_yth else ''}
    ✝️ &nbsp;YTH Camp
  </button>
</div>
""", unsafe_allow_html=True)

# helper: wrap in main padding
def _card(content, extra_style=''):
    return f'<div style="padding:0 1rem;{extra_style}">{content}</div>'

def sec_header(num, title, note=''):
    note_html = f'<span style="font-weight:400;text-transform:none;letter-spacing:0;font-size:.75rem;color:#aaa;margin-left:.4rem;">{note}</span>' if note else ''
    return st.markdown(f"""
<div class="sec-head {_sec_cls}" style="padding:0 1rem;margin:1.6rem 0 .7rem;">
  <div class="sec-num">{num}</div>
  <div class="sec-title">{title}{note_html}</div>
</div>""", unsafe_allow_html=True)

def param_card_open(title):
    st.markdown(f'<div class="param-card {_pcard}" style="margin:0 1rem .9rem;"><h4>{title}</h4>',
                unsafe_allow_html=True)

def param_card_close():
    st.markdown('</div>', unsafe_allow_html=True)

def upload_card_open(slot_class, label, badge_text, badge_opt, desc):
    opt_style = 'border-style:dashed;opacity:.88;' if badge_opt else ''
    badge_cls = 'opt' if badge_opt else ''
    st.markdown(f"""
<div class="upload-card {slot_class}" style="{opt_style}margin-bottom:.7rem;">
  <div class="upload-label">
    {label}
    <span class="badge {badge_cls}">{badge_text}</span>
  </div>
  <div class="upload-desc">{desc}</div>
""", unsafe_allow_html=True)

def upload_card_close():
    st.markdown('</div>', unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════════════════════
# STEP 1 — REGISTRATION FILE
# ══════════════════════════════════════════════════════════════════════════════
sec_header(1, "Registration File")

upload_card_open('reg', '① Registration', 'REQUIRED', False,
                 'Export from your registration platform — CSV or Excel. Must include First Name, Last Name, Grade, Gender.')
st.markdown('<div style="padding:0 1rem 1rem;">', unsafe_allow_html=True)
uploaded_reg = st.file_uploader("Registration file", type=['csv','xlsx'],
                                 label_visibility='collapsed', key='reg')
st.markdown('</div>', unsafe_allow_html=True)
upload_card_close()

if not uploaded_reg:
    st.markdown('</div>', unsafe_allow_html=True)  # close main offset
    st.stop()

with st.spinner("Reading registration file…"):
    try:
        df_raw = read_file(uploaded_reg)
        df_raw.columns = df_raw.columns.str.strip()
        friend_col_original = normalize_friend_col(df_raw)
        df_raw['_FullName'] = df_raw['First Name'].str.strip() + ' ' + df_raw['Last Name'].str.strip()
    except Exception as e:
        st.error(f"Could not read file: {e}"); st.stop()

has_grade     = 'Grade'  in df_raw.columns
has_gender    = 'Gender' in df_raw.columns
has_friend    = 'Friend Request' in df_raw.columns
has_community = 'Community Group' in df_raw.columns
detected_grades = sorted(df_raw['Grade'].dropna().astype(str).str.strip().unique().tolist()) \
                  if has_grade else []

st.markdown(f"""
<div style="padding:0 1rem .3rem;">
  <div class="ok-banner">
    ✅ <strong>{len(df_raw)} registrants loaded</strong>
    {'&nbsp;·&nbsp;Grades: ' + ', '.join(detected_grades) if detected_grades else ''}
    {'&nbsp;·&nbsp;Has gender' if has_gender else ''}
    {'&nbsp;·&nbsp;Has friend requests' if has_friend else ''}
    {f'&nbsp;·&nbsp;<em>Friend column renamed from "{friend_col_original}"</em>' if friend_col_original else ''}
  </div>
</div>
""", unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════════════════════
# STEP 2 — COMMUNITY / SUPPLEMENTAL FILES  (YTH only)
# ══════════════════════════════════════════════════════════════════════════════
df_community = None; df_leaders = None; df_prev = None

if camp_type == 'YTH':
    sec_header(2, "Community & Leader Files",
               "Upload what you have — each file improves group accuracy")

    # Row 1: HS Dashboard + Leaders side by side
    col_hs, col_ldr = st.columns(2)
    with col_hs:
        upload_card_open('cg', '② HS Community Dashboard', 'RECOMMENDED', False,
                         'Current-year YTH dashboard Excel. Grade sheets (Grade 9–12) with leader names.')
        st.markdown('<div style="padding:0 0 1rem;">', unsafe_allow_html=True)
        uploaded_community = st.file_uploader("HS Dashboard", type=['csv','xlsx','xls'],
                                               label_visibility='collapsed', key='community')
        st.markdown('</div>', unsafe_allow_html=True)
        upload_card_close()

    with col_ldr:
        upload_card_open('ldr', '③ Camp Leaders CSV', 'RECOMMENDED', False,
                         'Volunteer signup export. Used to map leaders to grades and genders.')
        st.markdown('<div style="padding:0 0 1rem;">', unsafe_allow_html=True)
        uploaded_leaders = st.file_uploader("Leaders CSV", type=['csv','xlsx'],
                                             label_visibility='collapsed', key='leaders')
        st.markdown('</div>', unsafe_allow_html=True)
        upload_card_close()

    # Row 2: MS Dashboard + Previous Groups
    col_ms, col_prev = st.columns(2)
    with col_ms:
        upload_card_open('supp', '④ MS Community Dashboard', 'OPTIONAL', True,
                         'Previous-year MS dashboard. 8th graders → 9th, 7th → 8th, etc. Grade promotion is automatic.')
        st.markdown('<div style="padding:0 0 1rem;">', unsafe_allow_html=True)
        uploaded_ms = st.file_uploader("MS Dashboard", type=['xlsx','xls'],
                                        label_visibility='collapsed', key='ms_community')
        st.markdown('</div>', unsafe_allow_html=True)
        upload_card_close()

    with col_prev:
        upload_card_open('supp', '⑤ Previous Camp Groups', 'OPTIONAL', True,
                         'Last year\'s camp group assignments. Used as tiebreaker when friend matches are ambiguous.')
        st.markdown('<div style="padding:0 0 1rem;">', unsafe_allow_html=True)
        uploaded_prev = st.file_uploader("Previous groups", type=['csv','xlsx'],
                                          label_visibility='collapsed', key='prev')
        st.markdown('</div>', unsafe_allow_html=True)
        upload_card_close()

    # ── Process HS dashboard ──────────────────────────────────────────────────
    if uploaded_community:
        try:
            df_community, parse_report = parse_community_dashboard(uploaded_community)
            if df_community is not None and len(df_community):
                report_lines = '<br>'.join(parse_report)
                st.markdown(f"""
<div style="padding:0 1rem;">
  <div class="ok-banner">✅ <strong>HS Dashboard</strong> — {len(df_community)} attendees,
  {df_community['Community Group'].nunique()} groups<br>
  <span style="font-size:.78rem;opacity:.75;">{report_lines}</span></div>
</div>""", unsafe_allow_html=True)
                df_community['_FullName'] = (df_community['First Name'].str.strip()
                                              + ' ' + df_community['Last Name'].str.strip())
                cg_map     = dict(zip(df_community['_FullName'].str.lower(), df_community['Community Group']))
                leader_map = dict(zip(df_community['_FullName'].str.lower(), df_community['Leader']))
                df_raw['Community Group'] = df_raw['_FullName'].str.lower().map(cg_map).fillna('')
                df_raw['Leader']          = df_raw['_FullName'].str.lower().map(leader_map).fillna('')
                has_community = True
                matched   = (df_raw['Community Group'] != '').sum()
                match_pct = round(100 * matched / len(df_raw)) if len(df_raw) else 0
                banner_cls = 'warn-banner' if match_pct < 70 else 'info-banner'
                st.markdown(f'<div style="padding:0 1rem;"><div class="{banner_cls}">'
                            f'{"⚠" if match_pct < 70 else "ℹ️"} '
                            f'<strong>{matched}/{len(df_raw)} names matched</strong> ({match_pct}%)'
                            f'</div></div>', unsafe_allow_html=True)
            else:
                uploaded_community.seek(0)
                df_community = read_file(uploaded_community)
                df_community.columns = df_community.columns.str.strip()
                if 'Community Group' in df_community.columns:
                    df_community['_FullName'] = (df_community['First Name'].str.strip()
                                                  + ' ' + df_community['Last Name'].str.strip())
                    cg_map = dict(zip(df_community['_FullName'].str.lower(), df_community['Community Group']))
                    df_raw['Community Group'] = df_raw['_FullName'].str.lower().map(cg_map).fillna('')
                    has_community = True
                    st.markdown('<div style="padding:0 1rem;"><div class="info-banner">'
                                'ℹ️ Community Group merged (flat CSV format).</div></div>',
                                unsafe_allow_html=True)
        except Exception as e:
            st.error(f"HS community file error: {e}")

    # ── Process MS dashboard ──────────────────────────────────────────────────
    if uploaded_ms:
        try:
            df_ms, ms_report = parse_ms_dashboard(uploaded_ms, promote_grades=True)
            if df_ms is not None and len(df_ms):
                report_lines = '<br>'.join(ms_report)
                df_ms['_FN'] = df_ms['First Name'].str.strip() + ' ' + df_ms['Last Name'].str.strip()
                hs_filled = df_raw.get('Community Group', pd.Series([''] * len(df_raw))) != ''
                ms_cg_m  = df_raw['_FullName'].str.lower().map(dict(zip(df_ms['_FN'].str.lower(), df_ms['Community Group'])))
                ms_ldr_m = df_raw['_FullName'].str.lower().map(dict(zip(df_ms['_FN'].str.lower(), df_ms['Leader'])))
                if 'Community Group' not in df_raw.columns: df_raw['Community Group'] = ''
                if 'Leader' not in df_raw.columns:          df_raw['Leader'] = ''
                df_raw.loc[~hs_filled & ms_cg_m.notna(), 'Community Group'] = ms_cg_m[~hs_filled & ms_cg_m.notna()]
                df_raw.loc[~hs_filled & ms_ldr_m.notna(), 'Leader']         = ms_ldr_m[~hs_filled & ms_ldr_m.notna()]
                has_community = True
                ms_added  = (~hs_filled & ms_cg_m.notna()).sum()
                tot_now   = (df_raw['Community Group'] != '').sum()
                tot_pct   = round(100 * tot_now / len(df_raw)) if len(df_raw) else 0
                st.markdown(f"""
<div style="padding:0 1rem;">
  <div class="ok-banner">✅ <strong>MS Dashboard</strong> — {len(df_ms)} attendees promoted by one grade,
  +{ms_added} new matches → combined coverage <strong>{tot_now}/{len(df_raw)} ({tot_pct}%)</strong><br>
  <span style="font-size:.78rem;opacity:.75;">{report_lines}</span></div>
</div>""", unsafe_allow_html=True)
            else:
                st.markdown(f'<div style="padding:0 1rem;"><div class="warn-banner">⚠ Could not parse MS dashboard.'
                            f'<br>{"<br>".join(ms_report)}</div></div>', unsafe_allow_html=True)
        except Exception as e:
            st.error(f"MS community file error: {e}")

    # ── Process leaders CSV ───────────────────────────────────────────────────
    if uploaded_leaders:
        try:
            df_leaders = read_file(uploaded_leaders)
            df_leaders.columns = df_leaders.columns.str.strip()
            st.markdown(f'<div style="padding:0 1rem;"><div class="ok-banner">'
                        f'✅ <strong>Leaders file</strong> — {len(df_leaders)} leaders loaded, '
                        f'columns: {", ".join(df_leaders.columns[:5].tolist())}</div></div>',
                        unsafe_allow_html=True)
        except Exception as e:
            st.error(f"Leaders file error: {e}")

    # ── Process previous groups ───────────────────────────────────────────────
    if uploaded_prev:
        try:
            df_prev = read_file(uploaded_prev)
            df_prev.columns = df_prev.columns.str.strip()
            st.markdown(f'<div style="padding:0 1rem;"><div class="ok-banner">'
                        f'✅ <strong>Previous groups</strong> — {len(df_prev)} records loaded'
                        f'</div></div>', unsafe_allow_html=True)
        except Exception as e:
            st.error(f"Previous groups file error: {e}")

else:
    # KIDS — no community files needed, just a placeholder variable
    uploaded_community = uploaded_ms = uploaded_leaders = uploaded_prev = None

# ══════════════════════════════════════════════════════════════════════════════
# STEP 3 — CONFIGURE PARAMETERS
# ══════════════════════════════════════════════════════════════════════════════
step_num = 3 if camp_type == 'YTH' else 2
sec_header(step_num, "Configure Parameters")

# ── Sorting splits ────────────────────────────────────────────────────────────
param_card_open("Sorting Splits")
col1, col2, col3 = st.columns(3)
with col1:
    separate_grade = st.toggle("Sort by Grade", value=has_grade, disabled=not has_grade,
                                help="Requires a 'Grade' column.")
with col2:
    separate_gender = st.toggle("Sort by Gender", value=False, disabled=not has_gender,
                                 help="Requires a 'Gender' column.")
with col3:
    if camp_type == 'YTH':
        separate_community = st.toggle("Sort by Community Group",
                                        value=has_community,
                                        help="Uses the Community Group column — upload a dashboard above to populate it.")
    else:
        separate_community = False
        st.write("")
param_card_close()

# ── Grade filter ──────────────────────────────────────────────────────────────
grade_filter = detected_grades
if has_grade and detected_grades:
    param_card_open("Grade Filter <span style='font-weight:400;text-transform:none;letter-spacing:0;font-size:.75rem;color:#aaa;'>(optional — leave all selected to include everyone)</span>")
    grade_filter = st.multiselect("Grades to include:", options=detected_grades, default=detected_grades)
    if not grade_filter:
        st.warning("No grades selected — all will be included.")
        grade_filter = detected_grades
    param_card_close()

# ── Group sizing priority ─────────────────────────────────────────────────────
param_card_open("Group Sizing Priority")
friend_mode_str = st.radio(
    "Sizing priority",
    options=['size', 'friends'],
    format_func=lambda x: (
        "🎯  Prioritize group size — fill groups to a target size, fitting friend pairs within the cap"
        if x == 'size' else
        "💚  Honor all friend requests — groups form around friend connections (size will vary)"
    ),
    label_visibility='collapsed',
    disabled=not has_friend,
)
friend_mode = (friend_mode_str == 'friends')

group_size = 8
if not friend_mode:
    group_size = st.slider("Target group size", min_value=4, max_value=24, value=8, step=1)
    if not has_friend:
        st.caption("⚠ No 'Friend Request' column detected — groups will be filled by grade/gender only.")
else:
    st.caption("Group size is determined by friend connections. Friendless kids are distributed evenly.")

camp_name = st.text_input(
    "Camp name  (used in the Excel header)",
    value=f"{'Rock Point YTH' if camp_type=='YTH' else 'Rock Point KIDS'} Camp 2026",
    max_chars=60
)
param_card_close()

# ══════════════════════════════════════════════════════════════════════════════
# GENERATE BUTTON
# ══════════════════════════════════════════════════════════════════════════════
st.markdown('<div style="padding:0 1rem;">', unsafe_allow_html=True)
generate = st.button("⚙️  Generate Sorted Groups", type="primary", use_container_width=True)
st.markdown('</div>', unsafe_allow_html=True)

if not generate:
    st.markdown('</div>', unsafe_allow_html=True)  # close sidebar offset
    st.stop()

# ══════════════════════════════════════════════════════════════════════════════
# PROCESS & SORT
# ══════════════════════════════════════════════════════════════════════════════
with st.spinner("Preparing data…"):
    df_work = df_raw.copy()
    if has_grade:
        df_work['Grade'] = df_work['Grade'].astype(str).str.strip()
        missing_mask = ~df_work['Grade'].isin(grade_filter)
    else:
        df_work['Grade'] = 'All'
        missing_mask = pd.Series([False]*len(df_work))

    df_missing = df_work[missing_mask].copy()
    df_clean   = df_work[~missing_mask].copy().reset_index(drop=True)

    if len(df_missing):
        names = ', '.join(df_missing['_FullName'].tolist()[:8])
        extra = f' … and {len(df_missing)-8} more' if len(df_missing) > 8 else ''
        st.markdown(f'<div style="padding:0 1rem;"><div class="warn-banner">⚠ '
                    f'<strong>{len(df_missing)} attendee(s)</strong> excluded (grade outside filter): '
                    f'{names}{extra}</div></div>', unsafe_allow_html=True)

with st.spinner("Sorting groups…"):
    params = dict(group_size=group_size, friend_mode=friend_mode, camp_type=camp_type,
                  camp_name=camp_name, separate_grade=separate_grade,
                  separate_gender=separate_gender, separate_community=separate_community)
    by_bucket = build_buckets(df_clean, separate_grade, separate_gender,
                               separate_community=(camp_type=='YTH' and separate_community))
    group_leaders = None
    if camp_type == 'YTH' and separate_community:
        all_groups, nmap, lmap, fmap, group_leaders = sort_yth_community(
            df_clean, by_bucket, group_size, friend_mode, df_leaders, df_prev)
    elif friend_mode:
        all_groups, nmap, lmap, fmap = sort_by_friends(df_clean, by_bucket)
    else:
        all_groups, nmap, lmap, fmap = sort_by_size(df_clean, by_bucket, group_size)
    compute_stats = make_compute_stats(df_clean, all_groups, nmap, lmap, fmap)

with st.spinner("Building Excel…"):
    excel_buf, total_kids, total_groups, n_missing, sat, total_req = \
        build_excel(df_clean, df_missing, all_groups, compute_stats, params, group_leaders)

# ══════════════════════════════════════════════════════════════════════════════
# RESULTS
# ══════════════════════════════════════════════════════════════════════════════
friend_pct = f"{round(100*sat/total_req)}%" if total_req else "N/A"
friend_cls = "good" if (friend_mode and total_req and sat > 0) else ""

st.markdown(f"""
<div style="padding:0 1rem;">
<div class="stats-row">
  <div class="stat-box">
    <div class="val">{len(df_raw)}</div><div class="lbl">Registered</div>
  </div>
  <div class="stat-box">
    <div class="val">{total_groups}</div><div class="lbl">Groups</div>
  </div>
  <div class="stat-box {friend_cls}">
    <div class="val">{friend_pct}</div><div class="lbl">Friend Req Met</div>
  </div>
  <div class="stat-box {'warn' if n_missing else ''}">
    <div class="val">{n_missing}</div><div class="lbl">Outside Filter</div>
  </div>
</div>
</div>
""", unsafe_allow_html=True)

# Bucket breakdown
seen_b = []; bucket_order_ui = []
for key, _ in all_groups:
    if key not in seen_b: seen_b.append(key); bucket_order_ui.append(key)

dots = ['#E25D32','#1A5276','#8AB738','#F3B73F','#6C3483','#1EBFE1',
        '#922B21','#2471A3','#0E6655','#7D6608','#4A235A','#1B4F72']
rows_html = ""
for bi, bucket in enumerate(bucket_order_ui):
    gs  = [(gd,g) for gd,g in all_groups if gd==bucket]
    nk  = sum(len(g) for _,g in gs)
    avg = f"{nk/len(gs):.1f}" if gs else '—'
    dot = dots[bi % len(dots)]
    rows_html += (f'<tr><td><span class="grade-dot" style="background:{dot}"></span>'
                  f'{bucket}</td><td>{nk}</td><td>{len(gs)}</td><td>{avg}</td></tr>')

st.markdown(f"""
<div style="padding:0 1rem 1rem;">
<table class="grade-table">
  <thead><tr><th>Bucket</th><th>Attendees</th><th>Groups</th><th>Avg Size</th></tr></thead>
  <tbody>{rows_html}</tbody>
</table>
</div>
""", unsafe_allow_html=True)

if camp_type == 'YTH' and group_leaders:
    assigned = sum(1 for cg, ldr in group_leaders.values() if ldr)
    st.markdown(f'<div style="padding:0 1rem;"><div class="ok-banner">✝️ '
                f'<strong>Community-first sort complete</strong> — '
                f'{assigned}/{total_groups} groups have an assigned leader from their community group.'
                f'</div></div>', unsafe_allow_html=True)

# Download
st.markdown('<div style="padding:0 1rem;">', unsafe_allow_html=True)
st.download_button(
    label="⬇  Download Sorted Groups (.xlsx)",
    data=excel_buf,
    file_name=f"{camp_name.lower().replace(' ','_')}_groups.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)
st.markdown('</div>', unsafe_allow_html=True)

# Close sidebar offset div
st.markdown('</div>', unsafe_allow_html=True)

def build_name_maps(df):
    name_map_full = {row['_FullName'].lower(): i for i, row in df.iterrows()}
    last_map = defaultdict(list); first_map = defaultdict(list)
    for i, row in df.iterrows():
        last_map[row['Last Name'].strip().lower()].append(i)
        first_map[row['First Name'].strip().lower()].append(i)
    return name_map_full, last_map, first_map

def find_idx(token, grade, df, nmap, lmap, fmap):
    t = token.lower().strip()
    if t in nmap: return nmap[t]
    if t in lmap:
        same = [c for c in lmap[t] if df.iloc[c].get('Grade','') == grade]
        if len(same) == 1: return same[0]
        if len(lmap[t]) == 1: return lmap[t][0]
    if t in fmap:
        same = [c for c in fmap[t] if df.iloc[c].get('Grade','') == grade]
        if len(same) == 1: return same[0]
        if len(fmap[t]) == 1: return fmap[t][0]
    words = token.split()
    for i in range(len(words)-1):
        cand = (words[i]+' '+words[i+1]).lower()
        if cand in nmap: return nmap[cand]
    return None

def get_friend_idxs(pidx, grade, df, nmap, lmap, fmap):
    req = df.iloc[pidx].get('Friend Request', '')
    return [fi for fi in (find_idx(t, grade, df, nmap, lmap, fmap)
                          for t in tokenize_request(req))
            if fi is not None and fi != pidx]

# ══════════════════════════════════════════════════════════════════════════════
# BUCKET BUILDER  (shared by both camp types)
# ══════════════════════════════════════════════════════════════════════════════
def build_buckets(df, separate_grade, separate_gender, separate_community=False):
    by_bucket = defaultdict(list)
    for i, row in df.iterrows():
        parts = []
        if separate_grade:
            parts.append(str(row.get('Grade','')).strip() or 'No Grade')
        if separate_gender:
            g = str(row.get('Gender','')).strip().lower()
            parts.append('Girls' if g in ('f','female','girl') else 'Boys')
        if separate_community:
            cg = str(row.get('Community Group','')).strip() or 'No Community'
            parts.append(cg)
        key = ' — '.join(parts) if parts else 'All Attendees'
        by_bucket[key].append(i)
    return by_bucket

# ══════════════════════════════════════════════════════════════════════════════
# SORTING ALGORITHMS
# ══════════════════════════════════════════════════════════════════════════════
def sort_by_size(df, by_bucket, group_size):
    nmap, lmap, fmap = build_name_maps(df)
    def gfi(pidx, grade): return get_friend_idxs(pidx, grade, df, nmap, lmap, fmap)
    def fs(pidx, grp, grade): return sum(1 for fi in gfi(pidx, grade) if fi in grp)

    all_groups = []; assigned = set()
    for bkey, members in by_bucket.items():
        grades = [df.iloc[i].get('Grade','') for i in members]
        grade = max(set(grades), key=grades.count)
        pool = sorted([i for i in members if i not in assigned], key=lambda i: -len(gfi(i, grade)))
        mset = set(members)
        for seed in pool:
            if seed in assigned: continue
            grp = {seed}; assigned.add(seed)
            for fi in gfi(seed, grade):
                if len(grp) >= group_size: break
                if fi not in assigned and fi in mset:
                    grp.add(fi); assigned.add(fi)
                    for ffi in gfi(fi, grade):
                        if len(grp) >= group_size: break
                        if ffi not in assigned and ffi in mset:
                            grp.add(ffi); assigned.add(ffi)
            remaining = sorted([i for i in members if i not in assigned],
                               key=lambda i: -fs(i, grp, grade))
            for p in remaining:
                if len(grp) >= group_size: break
                grp.add(p); assigned.add(p)
            all_groups.append((bkey, grp))
    # merge singletons
    for gi in range(len(all_groups)-1, -1, -1):
        key, grp = all_groups[gi]
        if len(grp) == 1:
            for pgi in range(gi-1, -1, -1):
                if all_groups[pgi][0] == key:
                    all_groups[pgi][1].update(grp); all_groups.pop(gi); break
    return all_groups, nmap, lmap, fmap

def sort_by_friends(df, by_bucket):
    nmap, lmap, fmap = build_name_maps(df)
    def gfi(pidx, grade): return get_friend_idxs(pidx, grade, df, nmap, lmap, fmap)

    all_groups = []
    for bkey, members in by_bucket.items():
        mset = set(members)
        grades = [df.iloc[i].get('Grade','') for i in members]
        grade = max(set(grades), key=grades.count)
        adj = defaultdict(set)
        for i in members:
            for fi in gfi(i, grade):
                if fi in mset: adj[i].add(fi); adj[fi].add(i)
        visited = set(); components = []
        for start in members:
            if start in visited or not adj[start]: continue
            comp = set(); queue = [start]
            while queue:
                node = queue.pop()
                if node in visited: continue
                visited.add(node); comp.add(node)
                queue.extend(adj[node] - visited)
            components.append(comp)
        friendless = [i for i in members if i not in visited]
        if not components: components.append(set(friendless)); friendless = []
        for fi in friendless: min(components, key=lambda c: len(c)).add(fi)
        for comp in components: all_groups.append((bkey, comp))
    return all_groups, nmap, lmap, fmap


def sort_yth_community(df, by_bucket, group_size, friend_mode,
                       df_leaders=None, df_prev=None):
    """
    YTH community-first sort — refined from manual group analysis.

    Key insight from comparing algorithm output to manual assignments:
    The manual process treats COMMUNITY GROUP = CAMP GROUP.
    Each community group (from the dashboard, ~5-17 kids) becomes one camp group,
    regardless of a hard size cap. Groups are NOT split to hit a target size.

    Sort priority:
    1. Grade + Gender (buckets — always enforced)
    2. Community Group — all members of the same CG stay together as one group.
       Multiple small CGs are merged only if both are under min_group_size (5).
    3. Friend Requests — used to pull unmatched kids into the right group.
       If a kid's friend is already in a CG group, that kid joins that group.
    4. Previous Camp Group — used as tiebreaker when friend match is ambiguous.
    5. Unmatched remainder — distributed evenly across existing groups by size.

    Leaders are assigned per community group from the leader embedded in the
    community group label (e.g. "10th Grade Girls — Cami").
    """
    nmap, lmap, fmap = build_name_maps(df)
    def gfi(pidx, grade): return get_friend_idxs(pidx, grade, df, nmap, lmap, fmap)

    # Build previous group lookup
    prev_lookup = {}
    if df_prev is not None:
        for _, row in df_prev.iterrows():
            fn = str(row.get('First Name','')).strip()
            ln = str(row.get('Last Name','')).strip()
            pg = str(row.get('Group', row.get('Camp Group', row.get('Group Name','')))).strip()
            if fn and ln and pg:
                prev_lookup[(fn+' '+ln).lower()] = pg

    all_groups   = []
    group_leaders = {}

    for bkey, members in by_bucket.items():
        grades = [df.iloc[i].get('Grade','') for i in members]
        grade  = max(set(grades), key=grades.count)
        member_set = set(members)

        # ── Step 1: Cluster members by community group ────────────────────────
        cg_clusters = defaultdict(list)   # cg_label -> [idx]
        no_cg       = []                  # members with no community group

        for i in members:
            cg = str(df.iloc[i].get('Community Group','')).strip()
            if cg:
                cg_clusters[cg].append(i)
            else:
                no_cg.append(i)

        # ── Step 2: Each CG = one group; merge tiny CGs (< min_size) ─────────
        MIN_MERGE = 4   # only merge CGs with fewer than 4 matched members
        # Cap at which we stop pulling unmatched into a group (loose, not hard)
        SOFT_CAP  = 18

        cg_groups  = []  # list of [cg_label, set_of_idxs]

        # Sort CGs by size descending so large ones anchor first
        for cg, members_in_cg in sorted(cg_clusters.items(),
                                         key=lambda x: -len(x[1])):
            cg_groups.append([cg, set(members_in_cg)])

        # Merge CGs that are too small to stand alone
        changed = True
        while changed:
            changed = False
            for i in range(len(cg_groups)-1, -1, -1):
                cg_label, grp = cg_groups[i]
                if len(grp) < MIN_MERGE and len(cg_groups) > 1:
                    # Prefer merging into a group whose leader is related
                    # (same leader first name match), otherwise smallest
                    best = None
                    ldr_part = cg_label.split(' — ')[-1].lower() if ' — ' in cg_label else ''
                    for j, (other_lbl, other_grp) in enumerate(cg_groups):
                        if j == i: continue
                        other_ldr = other_lbl.split(' — ')[-1].lower() if ' — ' in other_lbl else ''
                        # Prefer groups with overlapping leader first name token
                        if ldr_part and any(t in other_ldr for t in ldr_part.split()[:1]):
                            best = j; break
                    if best is None:
                        best = min((j for j in range(len(cg_groups)) if j != i),
                                   key=lambda j: len(cg_groups[j][1]))
                    cg_groups[best][1].update(grp)
                    cg_groups.pop(i)
                    changed = True
                    break

        # ── Step 3: Assign unmatched via friend request affinity ──────────────
        assigned = set(i for _, grp in cg_groups for i in grp)
        unmatched = [i for i in no_cg if i not in assigned]

        for i in unmatched:
            # Find which CG group contains the most of this person's friends
            friends  = gfi(i, grade)
            best_grp = None
            best_score = -1

            for gi, (cg_label, grp) in enumerate(cg_groups):
                score = sum(1 for fi in friends if fi in grp)
                # Tiebreak: previous camp group match
                if score == best_score and prev_lookup:
                    my_prev = prev_lookup.get(df.iloc[i]['_FullName'].lower(),'')
                    if my_prev and any(prev_lookup.get(df.iloc[j]['_FullName'].lower(),'') == my_prev
                                       for j in grp):
                        best_grp = gi
                if score > best_score:
                    best_score = score
                    best_grp = gi

            if best_grp is not None:
                cg_groups[best_grp][1].add(i)
            elif cg_groups:
                # No friend match — put in smallest group
                min(cg_groups, key=lambda x: len(x[1]))[1].add(i)
            else:
                cg_groups.append(['Unassigned', {i}])

        # ── Step 4: Emit groups, extract leader from CG label ─────────────────
        gi_offset = len(all_groups)
        for gi, (cg_label, grp) in enumerate(cg_groups):
            all_groups.append((bkey, grp))
            # Leader = text after ' — ' in the community group label
            leader = cg_label.split(' — ')[-1].strip() if ' — ' in cg_label else ''
            group_leaders[gi_offset + gi] = (cg_label, leader)

    # Build idx_to_group for compute_stats
    idx_to_group = {}
    for _, grp in all_groups:
        for i in grp: idx_to_group[i] = grp

    def compute_stats_inner(pidx, grade):
        friends = gfi(pidx, grade)
        grp     = idx_to_group.get(pidx, set())
        return len(friends), sum(1 for fi in friends if fi in grp)

    return all_groups, nmap, lmap, fmap, group_leaders


def make_compute_stats(df, all_groups, nmap, lmap, fmap):
    idx_to_group = {}
    for _, grp in all_groups:
        for i in grp: idx_to_group[i] = grp
    def gfi(pidx, grade): return get_friend_idxs(pidx, grade, df, nmap, lmap, fmap)
    def compute_stats(pidx, grade):
        friends = gfi(pidx, grade); grp = idx_to_group.get(pidx, set())
        return len(friends), sum(1 for fi in friends if fi in grp)
    return compute_stats

# ══════════════════════════════════════════════════════════════════════════════
# EXCEL BUILDER
# ══════════════════════════════════════════════════════════════════════════════
def build_excel(df, df_missing, all_groups, compute_stats, params,
                group_leaders=None):
    group_size    = params.get('group_size', 8)
    friend_mode   = params.get('friend_mode', False)
    camp_type     = params.get('camp_type', 'KIDS')
    camp_name     = params.get('camp_name', 'Camp 2026') or 'Camp 2026'
    separate_grade   = params.get('separate_grade', False)
    separate_gender  = params.get('separate_gender', False)
    separate_community = params.get('separate_community', False)

    KEEP_COLS = [c for c in KEEP_COLS_PREF if c in df.columns]
    HELPER_COLS = ['Friends_Requested', 'Friends_Met']
    if group_leaders:
        HELPER_COLS += ['Community_Group', 'Assigned_Leader']

    total_kids   = sum(len(g) for _,g in all_groups)
    total_groups = len(all_groups)
    sat = total_req = 0
    for key, grp in all_groups:
        grade_lbl = df.iloc[list(grp)[0]].get('Grade','')
        for i in grp:
            made, met = compute_stats(i, grade_lbl)
            total_req += made; sat += met

    seen = []; bucket_order = []
    for key, _ in all_groups:
        if key not in seen: seen.append(key); bucket_order.append(key)
    color_map = {k: bucket_color(i) for i, k in enumerate(bucket_order)}

    wb = openpyxl.Workbook(); wb.remove(wb.active)
    thin = Side(style='thin', color='D0D8F0')
    def card_border(): return Border(top=thin, left=thin, right=thin, bottom=thin)

    def write_header(ws, headers, color, extra_headers=None, extra_color=None, row=1):
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=row, column=ci, value=h)
            c.font = Font(bold=True, color=WHITE, name='Arial', size=10)
            c.fill = PatternFill('solid', start_color=color)
            c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        if extra_headers:
            for ei, h in enumerate(extra_headers, len(headers)+1):
                c = ws.cell(row=row, column=ei, value=h)
                c.font = Font(bold=True, color=WHITE, name='Arial', size=10)
                c.fill = PatternFill('solid', start_color=extra_color or color)
                c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.row_dimensions[row].height = 28

    def write_data_row(ws, ri, vals, fill, extra_vals=None, extra_fill=None):
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=ri, column=ci, value=v)
            c.font = Font(name='Arial', size=10)
            c.fill = PatternFill('solid', start_color=fill)
            c.alignment = Alignment(horizontal='left', vertical='center')
        if extra_vals:
            for ei, v in enumerate(extra_vals, len(vals)+1):
                c = ws.cell(row=ri, column=ei, value=v)
                c.font = Font(name='Arial', size=10)
                c.fill = PatternFill('solid', start_color=extra_fill or fill)
                c.alignment = Alignment(horizontal='center', vertical='center')

    def write_bar(ws, ri, ncols):
        for ci in range(1, ncols+1):
            c = ws.cell(row=ri, column=ci, value='')
            c.fill = PatternFill('solid', start_color=BLACK)
        ws.row_dimensions[ri].height = 6

    def set_widths(ws, headers, all_vals):
        for ci, h in enumerate(headers, 1):
            mx = len(h)
            for row in all_vals:
                if ci-1 < len(row): mx = max(mx, len(str(row[ci-1])))
            ws.column_dimensions[get_column_letter(ci)].width = min(mx+3, 32)

    def merge_style(ws, r1, c1, r2, c2, value, bold=False, size=11,
                    color='1A1A2E', bg=None, align='left'):
        ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
        c = ws.cell(row=r1, column=c1, value=value)
        c.font = Font(name='Arial', bold=bold, size=size, color=color)
        c.alignment = Alignment(horizontal=align, vertical='center')
        if bg: c.fill = PatternFill('solid', start_color=bg)
        return c

    camp_icon = '✝️' if camp_type == 'YTH' else '🏕️'
    mode_label = 'Friend-First' if friend_mode else f'Size-Targeted (≈{group_size})'
    split_parts = (['Grade'] if separate_grade else []) + \
                  (['Gender'] if separate_gender else []) + \
                  (['Community'] if separate_community else [])
    split_label = ' + '.join(split_parts) if split_parts else 'None'

    # ── Dashboard ──────────────────────────────────────────────────────────────
    ws_dash = wb.create_sheet(title='Dashboard', index=0)
    ws_dash.sheet_view.showGridLines = False
    for col, w in [('A',3),('B',28),('C',16),('D',16),('E',16),('F',16),('G',16),('H',3)]:
        ws_dash.column_dimensions[col].width = w
    for r in range(1, 400): ws_dash.row_dimensions[r].height = 18
    for r, h in [(1,10),(2,46),(3,18),(4,10),(5,16),(6,52),(7,14),(8,16),(9,26)]:
        ws_dash.row_dimensions[r].height = h
    for r in range(10, 10+len(bucket_order)+3): ws_dash.row_dimensions[r].height = 22
    roster_start = 10 + len(bucket_order) + 4
    ws_dash.row_dimensions[roster_start-2].height = 14
    ws_dash.row_dimensions[roster_start-1].height = 16
    ws_dash.row_dimensions[roster_start].height = 26
    for r in range(roster_start+1, 400): ws_dash.row_dimensions[r].height = 19

    merge_style(ws_dash, 2, 2, 2, 7,
                f'{camp_icon}  {camp_name} ({camp_type}) — Group Sorter',
                bold=True, size=18, color=WHITE, bg='1A1A2E', align='left')
    merge_style(ws_dash, 3, 2, 3, 7,
                f'Mode: {mode_label}   |   Split by: {split_label}   |   Attendees: {len(df)+len(df_missing)}',
                bold=False, size=9, color='888888', bg='F8F9FC', align='left')

    sum_total_cols = 2 + len(KEEP_COLS)
    req_col = get_column_letter(sum_total_cols+1)
    met_col = get_column_letter(sum_total_cols+2)
    friend_formula = (f'=IFERROR(TEXT(SUM(Summary!{met_col}:{met_col})'
                      f'/SUM(Summary!{req_col}:{req_col}),"0%"),"N/A")')
    fkpi_bg = 'EAF7EA' if friend_mode else 'F0F4FF'
    fkpi_fg = '2E7D32' if friend_mode else '1A1A2E'

    kpis = [
        ('Total Registered', len(df)+len(df_missing), 'F0F4FF','1A1A2E',False),
        ('Total Groups',     total_groups,             'F0F4FF','1A1A2E',False),
        ('Avg Group Size',   f"{total_kids/total_groups:.1f}", 'F0F4FF','1A1A2E',False),
        ('Friend Req Met',   friend_formula,           fkpi_bg, fkpi_fg, True),
        ('Outside Filter',   len(df_missing),          'FFF3CD','B8600A',False),
    ]
    for idx, (label, value, bg, fg, is_f) in enumerate(kpis):
        col = 2+idx
        for r, v, sz, bold in [(5,label,9,False),(6,value,22,True)]:
            c = ws_dash.cell(row=r, column=col, value=v)
            c.font = Font(name='Arial', bold=bold, size=sz, color='666666' if r==5 else fg)
            c.fill = PatternFill('solid', start_color=bg)
            c.alignment = Alignment(horizontal='center', vertical='bottom' if r==5 else 'center')
            c.border = card_border()
            if is_f and r==5:
                c.font = Font(name='Arial', bold=False, size=9, color='666666', italic=True)
                c.value = 'Friend Req Met ƒ'

    merge_style(ws_dash, 8, 2, 8, 7, 'ATTENDEES BY GROUP', bold=True, size=9, color='888888', bg=WHITE)
    bhdrs = ['Group','Attendees','Groups','Avg Size',
             'Largest' if friend_mode else 'Full Groups',
             'Smallest' if friend_mode else 'Partial']
    for ci, h in zip([2,3,4,5,6,7], bhdrs):
        c = ws_dash.cell(row=9, column=ci, value=h)
        c.font = Font(name='Arial', bold=True, size=10, color=WHITE)
        c.fill = PatternFill('solid', start_color='1A1A2E')
        c.alignment = Alignment(horizontal='center' if ci>2 else 'left', vertical='center')
    shades = ['F9EBEA','EBF5FB','E8F8F5','FEF9E7','F5EEF8','FDFEFE',
              'EAFAF1','FEF5E7','EBF5FB','F9EBEA']
    for ri, bucket in enumerate(bucket_order):
        rn = 10+ri
        gs = [(gd,g) for gd,g in all_groups if gd==bucket]
        nk = sum(len(g) for _,g in gs); ng = len(gs)
        avg = nk/ng if ng else 0
        v5 = max((len(g) for _,g in gs), default=0) if friend_mode \
             else sum(1 for _,g in gs if len(g) >= group_size)
        v6 = min((len(g) for _,g in gs), default=0) if friend_mode else ng-v5
        chex = color_map.get(bucket, '888888')
        for ci, v in zip([2,3,4,5,6,7], [bucket, nk, ng, f"{avg:.1f}", v5, v6]):
            c = ws_dash.cell(row=rn, column=ci, value=v)
            c.font = Font(name='Arial', bold=(ci==2), size=10,
                          color=chex if ci==2 else '1A1A2E')
            c.fill = PatternFill('solid', start_color=shades[ri % len(shades)])
            c.alignment = Alignment(horizontal='center' if ci>2 else 'left', vertical='center')
    rn = 10+len(bucket_order)
    for ci, v in zip([2,3,4,5,6,7], ['⚠ Outside Grade Filter', len(df_missing),'—','—','—','—']):
        c = ws_dash.cell(row=rn, column=ci, value=v)
        c.font = Font(name='Arial', bold=(ci==2), size=10, color='B8600A' if ci==2 else '1A1A2E')
        c.fill = PatternFill('solid', start_color='FFF3CD')
        c.alignment = Alignment(horizontal='center' if ci>2 else 'left', vertical='center')

    # Group roster
    roster_lbl_r = roster_start - 1
    merge_style(ws_dash, roster_lbl_r, 2, roster_lbl_r, 7, 'GROUP ROSTER',
                bold=True, size=9, color='888888', bg=WHITE)
    roster_hdrs = ['Group #', 'Bucket', 'Members', 'Status', 'Leader', 'Community Grp']
    for ci, h in zip([2,3,4,5,6,7], roster_hdrs):
        c = ws_dash.cell(row=roster_start, column=ci, value=h)
        c.font = Font(name='Arial', bold=True, size=10, color=WHITE)
        c.fill = PatternFill('solid', start_color='1A1A2E')
        c.alignment = Alignment(horizontal='center' if ci>2 else 'left', vertical='center')
    ws_dash.row_dimensions[roster_start].height = 26
    gnum = 1
    for bucket in bucket_order:
        chex = color_map.get(bucket, '888888')
        for gi_local, (gd, grp) in enumerate(
                [(gd,g) for gd,g in all_groups if gd==bucket]):
            global_gi = next(i for i,(k,g) in enumerate(all_groups)
                             if k==bucket and g is grp)
            rn = roster_start+gnum; n = len(grp)
            status = f'{n} kids' if friend_mode else (
                'Full' if n >= group_size else f'Open ({group_size-n} spot{"s" if group_size-n>1 else ""})')
            sc = '2E7D32' if (friend_mode or n >= group_size) else 'B8600A'
            shade = 'F7F7F7' if gnum%2==0 else WHITE
            leader_info = group_leaders.get(global_gi, ('','')) if group_leaders else ('','')
            cg_name, leader_name = leader_info
            for ci, v in zip([2,3,4,5,6,7],
                              [f"Group {gnum}", bucket, n, status, leader_name, cg_name]):
                c = ws_dash.cell(row=rn, column=ci, value=v)
                c.fill = PatternFill('solid', start_color=shade)
                c.font = Font(name='Arial', size=10,
                              color=chex if ci==3 else (sc if ci==5 else '1A1A2E'),
                              bold=(ci==3))
                c.alignment = Alignment(horizontal='center' if ci==4 else 'left', vertical='center')
            gnum += 1

    # ── Summary ────────────────────────────────────────────────────────────────
    ws_sum = wb.create_sheet(title='Summary')
    sum_hdrs = ['Group','Bucket'] + KEEP_COLS
    write_header(ws_sum, [h.strip() for h in sum_hdrs], '2E4057',
                 extra_headers=HELPER_COLS, extra_color=HELPER_HDR)
    c = ws_sum.cell(row=2, column=sum_total_cols+1, value='# friend names listed')
    c.font = Font(name='Arial', size=8, italic=True, color='555555')
    c.fill = PatternFill('solid', start_color='D6EAF8')
    c.alignment = Alignment(horizontal='center', vertical='center')
    c = ws_sum.cell(row=2, column=sum_total_cols+2, value='# placed in same group')
    c.font = Font(name='Arial', size=8, italic=True, color='555555')
    c.fill = PatternFill('solid', start_color='D6EAF8')
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws_sum.row_dimensions[2].height = 14

    cur = 3; all_sv = []; gnum = 1; prev_bucket = None
    for bucket in bucket_order:
        for gi_global, (gd, grp) in enumerate(all_groups):
            if gd != bucket: continue
            if prev_bucket and bucket != prev_bucket:
                write_bar(ws_sum, cur, sum_total_cols+len(HELPER_COLS)); cur += 1
            prev_bucket = bucket
            grade_lbl = df.iloc[list(grp)[0]].get('Grade','')
            members = sorted(grp, key=lambda i: df.iloc[i]['Last Name'])
            leader_info = group_leaders.get(gi_global, ('','')) if group_leaders else ('','')
            cg_name, leader_name = leader_info
            for idx, i in enumerate(members):
                row = df.iloc[i]
                made, met = compute_stats(i, grade_lbl)
                has_req = made > 0; unmet = has_req and met == 0
                vals = [f'Group {gnum}', bucket] + [get_val(row, c2) for c2 in KEEP_COLS]
                fill = 'FFFDE7' if unmet else ('F2F2F2' if idx%2==0 else WHITE)
                extra = [made, met]
                if group_leaders: extra += [cg_name, leader_name]
                write_data_row(ws_sum, cur, vals, fill, extra_vals=extra, extra_fill=HELPER_BG)
                all_sv.append(vals+extra); cur += 1
            write_bar(ws_sum, cur, sum_total_cols+len(HELPER_COLS)); cur += 1
            gnum += 1
    set_widths(ws_sum, [h.strip() for h in sum_hdrs]+HELPER_COLS, all_sv)
    ws_sum.column_dimensions[req_col].width = 20
    ws_sum.column_dimensions[met_col].width = 20

    # ── Per-bucket sheets ──────────────────────────────────────────────────────
    for bi, bucket in enumerate(bucket_order):
        gs = [(gd,g) for gd,g in all_groups if gd==bucket]
        if not gs: continue
        ws = wb.create_sheet(title=bucket[:31])
        ws.sheet_view.showGridLines = False
        write_header(ws, [h.strip() for h in KEEP_COLS], color_map.get(bucket,'888888'))
        cur = 2; all_gv = []
        for _, grp in gs:
            members = sorted(grp, key=lambda i: df.iloc[i]['Last Name'])
            for idx, i in enumerate(members):
                row = df.iloc[i]
                vals = [get_val(row, c2) for c2 in KEEP_COLS]
                write_data_row(ws, cur, vals, 'F2F2F2' if idx%2==0 else WHITE)
                all_gv.append(vals); cur += 1
            write_bar(ws, cur, len(KEEP_COLS)); cur += 1
        set_widths(ws, [h.strip() for h in KEEP_COLS], all_gv)

    # ── Leaders sheet (YTH only) ───────────────────────────────────────────────
    if group_leaders:
        ws_ldr = wb.create_sheet(title='Leaders')
        ws_ldr.sheet_view.showGridLines = False
        ldr_hdrs = ['Group #','Bucket','Community Group','Assigned Leader','# Members']
        write_header(ws_ldr, ldr_hdrs, '1A5276')
        ldr_data = []
        gnum = 1
        for bucket in bucket_order:
            for gi_global, (gd, grp) in enumerate(all_groups):
                if gd != bucket: continue
                cg_name, leader_name = group_leaders.get(gi_global, ('',''))
                row_vals = [f'Group {gnum}', bucket, cg_name, leader_name, len(grp)]
                ldr_data.append(row_vals)
                shade = 'F2F2F2' if gnum%2==0 else WHITE
                write_data_row(ws_ldr, gnum+1, row_vals, shade)
                ws_ldr.row_dimensions[gnum+1].height = 20
                gnum += 1
        set_widths(ws_ldr, ldr_hdrs, ldr_data)

    # ── Missing tab ────────────────────────────────────────────────────────────
    if len(df_missing):
        ws_miss = wb.create_sheet(title='Outside Filter')
        ws_miss.sheet_view.showGridLines = False
        ws_miss.merge_cells('A1:P1')
        b = ws_miss.cell(row=1, column=1,
            value='⚠  Grade outside filter or unrecognized — assign manually and re-run')
        b.font = Font(name='Arial', bold=True, size=11, color='7D4E00')
        b.fill = PatternFill('solid', start_color='FFF3CD')
        b.alignment = Alignment(horizontal='left', vertical='center')
        ws_miss.row_dimensions[1].height = 28
        all_cols = [c for c in df_missing.columns if not c.startswith('_')]
        write_header(ws_miss, all_cols, 'B8600A', row=2)
        for ri, (_, row) in enumerate(df_missing.iterrows(), 3):
            for ci, col in enumerate(all_cols, 1):
                c = ws_miss.cell(row=ri, column=ci, value=clean(row.get(col,'')))
                c.font = Font(name='Arial', size=10)
                c.fill = PatternFill('solid', start_color='FFFBF0')
                c.alignment = Alignment(horizontal='left', vertical='center')
            ws_miss.row_dimensions[ri].height = 20
        for ci, col in enumerate(all_cols, 1):
            vals = [str(clean(row.get(col,''))) for _, row in df_missing.iterrows()]
            mx = max(len(col), max((len(v) for v in vals), default=0))
            ws_miss.column_dimensions[get_column_letter(ci)].width = min(mx+3, 35)

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf, total_kids, total_groups, len(df_missing), sat, total_req

# ══════════════════════════════════════════════════════════════════════════════
# HELPERS
# ══════════════════════════════════════════════════════════════════════════════
def clean(val):
    if pd.isna(val) or str(val).strip().lower() in ('nan','n/a','na',''): return ''
    return str(val).strip()
def get_val(row, col):
    try: return clean(row.get(col,''))
    except: return ''

def normalize_friend_col(df):
    """Rename any known friend-request column variant to 'Friend Request' in-place."""
    for variant in FRIEND_REQUEST_VARIANTS:
        if variant in df.columns and variant != 'Friend Request':
            df.rename(columns={variant: 'Friend Request'}, inplace=True)
            return variant   # return original name so we can show it in the UI
    return None

def read_file(f):
    if f is None: return None
    name = f.name.lower()
    if name.endswith('.csv'):   return pd.read_csv(f)
    if name.endswith('.xlsx'):  return pd.read_excel(f, engine='openpyxl')
    if name.endswith('.xls'):   return pd.read_excel(f)
    return None

# ── Community Dashboard parser ─────────────────────────────────────────────────
# Handles the YTH Dashboard Excel format where each sheet is a grade
# (e.g. "Grade 12", "12th Grade", "Grade 12 Boys") and columns represent
# community groups led by a leader (row 2), with attendee names in rows 4+.
GRADE_SHEET_PATTERN = re.compile(r'\b(\d{1,2})\b')
GENDER_KEYWORDS = {
    'girl': 'Girl', 'girls': 'Girl', 'female': 'Girl',
    'guy':  'Boy',  'guys':  'Boy',  'male':   'Boy',
    'boy':  'Boy',  'boys':  'Boy',  'men':    'Boy', 'man': 'Boy',
}
ORDINALS = {'9':'9th','10':'10th','11':'11th','12':'12th',
            '6':'6th','7':'7th','8':'8th','5':'5th','4':'4th'}

def _is_grade_sheet(sheet_name):
    """Return grade number string if this sheet represents a grade, else None."""
    m = GRADE_SHEET_PATTERN.search(sheet_name)
    if not m: return None
    num = m.group(1)
    if int(num) < 4 or int(num) > 12: return None
    # Exclude obviously non-grade sheets by checking for grade-related words
    name_lower = sheet_name.lower()
    grade_words = ('grade','gr','th','nd','rd','st')
    if any(w in name_lower for w in grade_words): return num
    # Also accept pure number sheets like "12" or "9"
    if sheet_name.strip() == num: return num
    return None

def _detect_gender_sections(header_row):
    """Return dict of col_index -> 'Boy'|'Girl' from the header row."""
    col_gender = {}
    current = None
    for ci, val in enumerate(header_row):
        if val and isinstance(val, str):
            lower = val.lower()
            for kw, g in GENDER_KEYWORDS.items():
                if kw in lower:
                    current = g
                    break
        if current:
            col_gender[ci] = current
    return col_gender

def parse_community_dashboard(f):
    """
    Parse a YTH community dashboard Excel into a DataFrame with columns:
      First Name, Last Name, Grade, Gender, Community Group, Leader
    Supports sheet name patterns: "Grade 12", "12th Grade", "Grade 12 Boys", etc.
    Returns (df_records, parse_report) where parse_report lists what was found.
    """
    try:
        wb = load_workbook(f, read_only=True)
    except Exception as e:
        return None, [f"Could not open file: {e}"]

    all_records = []
    parse_report = []
    skipped_names = []

    for sheet_name in wb.sheetnames:
        grade_num = _is_grade_sheet(sheet_name)
        if grade_num is None:
            continue

        ws = wb[sheet_name]
        all_rows = list(ws.iter_rows(values_only=True))
        if len(all_rows) < 4:
            parse_report.append(f"  ⚠ {sheet_name}: fewer than 4 rows, skipped")
            continue

        header_row = all_rows[0]   # gender section headers
        leader_row = all_rows[1]   # leader names per column
        # all_rows[2] = rooms — skip
        data_rows  = all_rows[3:]

        col_gender = _detect_gender_sections(header_row)
        col_leader = {}
        for ci, val in enumerate(leader_row):
            if val and isinstance(val, str):
                v = val.strip()
                if v and v not in ('TOTAL',) and not v.startswith('='):
                    col_leader[ci] = v

        grade_label = f"{ORDINALS.get(grade_num, grade_num+'th')} Grade"

        # If sheet name itself has a gender (e.g. "Grade 12 Boys"), apply globally
        sheet_gender_override = None
        for kw, g in GENDER_KEYWORDS.items():
            if kw in sheet_name.lower():
                sheet_gender_override = g
                break

        col_group = {}
        for ci, leader in col_leader.items():
            gender = sheet_gender_override or col_gender.get(ci, 'Unknown')
            col_group[ci] = f"{grade_label} {gender}s — {leader}"

        sheet_records = 0
        for row in data_rows:
            for ci, val in enumerate(row):
                if not val or not isinstance(val, str): continue
                val = val.strip()
                if not val or val.upper() in ('TOTAL','N/A') or val.startswith('='): continue
                parts = val.split()
                if len(parts) < 2:
                    skipped_names.append(f"{sheet_name} col {ci}: '{val}' (no last name)")
                    continue
                first = parts[0].strip()
                last  = ' '.join(parts[1:]).strip()
                gender = sheet_gender_override or col_gender.get(ci, 'Unknown')
                leader = col_leader.get(ci, '')
                group  = col_group.get(ci, f"{grade_label} {gender}s")
                all_records.append({
                    'First Name':      first,
                    'Last Name':       last,
                    'Grade':           grade_label,
                    'Gender':          'Male' if gender == 'Boy' else 'Female',
                    'Community Group': group,
                    'Leader':          leader,
                })
                sheet_records += 1

        parse_report.append(f"  ✅ {sheet_name} → Grade {grade_num} ({sheet_records} attendees, {len(col_leader)} groups)")

    if skipped_names:
        parse_report.append(f"  ⚠ {len(skipped_names)} name(s) skipped (single-word, no last name):")
        for s in skipped_names[:5]:
            parse_report.append(f"     {s}")
        if len(skipped_names) > 5:
            parse_report.append(f"     … and {len(skipped_names)-5} more")

    if not all_records:
        return None, parse_report + ["❌ No attendee records found. Check sheet names contain a grade number (e.g. 'Grade 9', '10th Grade')."]

    df = pd.DataFrame(all_records)
    return df, parse_report

# ── Grade promotion map (dashboard grade → camp grade) ────────────────────────
# Used when cross-referencing MS dashboard: 8th graders become 9th at camp, etc.
GRADE_PROMOTION_MAP = {
    '6': '7th', '7': '8th', '8': '9th',
    '9': '10th', '10': '11th', '11': '12th',
    '12': None,  # graduated — exclude
}

def _patch_xlsx_stylesheet(file_obj):
    """
    Some Excel files have invalid font family values (>14) that crash openpyxl.
    Patch the stylesheet XML in-memory and return a BytesIO of the fixed file.
    """
    import zipfile, re as _re, io as _io
    data = file_obj.read() if hasattr(file_obj, 'read') else open(file_obj,'rb').read()
    buf = _io.BytesIO()
    with zipfile.ZipFile(_io.BytesIO(data), 'r') as zin:
        with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                raw = zin.read(item.filename)
                if item.filename == 'xl/styles.xml':
                    text = raw.decode('utf-8', errors='replace')
                    text = _re.sub(
                        r'<family val="(\d+)"/>',
                        lambda m: f'<family val="{min(int(m.group(1)), 14)}"/>',
                        text
                    )
                    raw = text.encode('utf-8')
                zout.writestr(item, raw)
    buf.seek(0)
    return buf

def parse_ms_dashboard(f, promote_grades=True):
    """
    Parse a Middle School community dashboard Excel.

    Layout per sheet (e.g. '8th Girls', '7th Boys', '6th Girls'):
      Row 1:   grade + gender title  ("8th Grade Girls")
      Row 2:   empty
      Rows 3–5: leader names — one full name per 3-col group (col 0, 3, 6 …)
               may span 1–3 rows; some groups have multiple co-leaders
      Row ~6:  empty separator
      Rows 6+: attendees as  First | Last | _ | First | Last | _ | …
               each community group occupies a 3-column band

    When promote_grades=True, grades are bumped by one year
    (8th → 9th, 7th → 8th, 6th → 7th) to reflect promotion to camp year.
    12th graders are excluded (graduated).

    Returns (df_records, parse_report).
    """
    # Try to load; patch stylesheet if needed
    try:
        fixed = _patch_xlsx_stylesheet(f)
        wb = load_workbook(fixed, read_only=True)
    except Exception as e:
        return None, [f"Could not open file: {e}"]

    # Identify grade sheets: name must match "Nth Girls/Boys/Guys" or "Nth Grade X"
    MS_SHEET_RE = re.compile(r'(\d{1,2})(?:st|nd|rd|th)?', re.IGNORECASE)
    MS_GENDER   = {'girl':'Girl','girls':'Girl','female':'Girl',
                   'guy':'Boy','guys':'Boy','boy':'Boy','boys':'Boy','male':'Boy'}

    all_records  = []
    parse_report = []
    skipped      = []

    for sheet_name in wb.sheetnames:
        m = MS_SHEET_RE.search(sheet_name)
        if not m: continue
        grade_num = m.group(1)
        if int(grade_num) < 4 or int(grade_num) > 12: continue

        # Must have a gender or grade keyword to qualify as a grade sheet
        name_lower = sheet_name.lower()
        has_gender_kw = any(k in name_lower for k in MS_GENDER)
        has_grade_kw  = any(k in name_lower for k in ('grade','gr','th','nd','rd','st'))
        if not (has_gender_kw or has_grade_kw): continue

        # Determine gender from sheet name
        sheet_gender = 'Unknown'
        for kw, g in MS_GENDER.items():
            if kw in name_lower:
                sheet_gender = g; break

        # Grade label for this sheet
        orig_label = f"{ORDINALS.get(grade_num, grade_num+'th')} Grade"

        # Promoted camp grade
        if promote_grades:
            promoted_num = GRADE_PROMOTION_MAP.get(grade_num)
            if promoted_num is None:
                parse_report.append(f"  ⏭ {sheet_name} — Grade {grade_num} graduated, skipped")
                continue
            camp_grade = f"{promoted_num} Grade"
        else:
            camp_grade = orig_label

        gender_label = 'Girls' if sheet_gender == 'Girl' else 'Boys'

        ws       = wb[sheet_name]
        all_rows = list(ws.iter_rows(values_only=True))
        if len(all_rows) < 4: continue

        # Auto-detect column offset: find the first column that has data
        # Some sheets start at col 0, others at col 1
        col_offset = 0
        for row in all_rows[1:6]:  # check rows 2-6
            for ci, val in enumerate(row):
                if val and isinstance(val, str) and val.strip() and 'NO MAN' not in val.upper():
                    col_offset = ci % 3  # offset is 0 if ci=0,3,6... or 1 if ci=1,4,7...
                    break
            if col_offset > 0: break

        # Find where data rows start: first row where col[offset] AND col[offset+1] both have strings
        data_start = None
        leader_rows_raw = []
        for ri, row in enumerate(all_rows):
            if ri == 0: continue  # skip title row
            # Check if any 3-col group (starting at col_offset) has both first AND last name
            is_data = any(
                ci < len(row) and row[ci] and isinstance(row[ci], str) and str(row[ci]).strip() and
                ci+1 < len(row) and row[ci+1] and isinstance(row[ci+1], str) and str(row[ci+1]).strip()
                for ci in range(col_offset, len(row)-1, 3)
            )
            if is_data:
                data_start = ri; break
            # Collect as potential leader row if it has content at offset positions
            has_content = any(
                ci < len(row) and row[ci] and isinstance(row[ci], str)
                and str(row[ci]).strip() and 'NO MAN' not in str(row[ci]).upper()
                for ci in range(col_offset, len(row), 3)
            )
            if has_content:
                leader_rows_raw.append(row)

        if data_start is None: continue

        # Build col_group -> leader name(s) using col_offset
        max_col    = max((len(r) for r in all_rows[data_start:data_start+5]), default=0)
        num_groups = max(1, (max_col - col_offset + 2) // 3)
        group_ldrs: dict = {}  # gi -> list[str]

        for row in leader_rows_raw:
            for gi in range(num_groups):
                ci  = col_offset + gi * 3
                val = row[ci] if ci < len(row) else None
                if not val or not isinstance(val, str): continue
                v = val.strip()
                if not v or 'NO MAN' in v.upper(): continue
                group_ldrs.setdefault(gi, [])
                if v not in group_ldrs[gi]:
                    group_ldrs[gi].append(v)

        def group_label(gi):
            ldrs = group_ldrs.get(gi, [])
            ldr_str = ' & '.join(ldrs) if ldrs else 'Unassigned'
            return f"{camp_grade} {gender_label} — {ldr_str}"

        # Parse attendee rows using col_offset
        sheet_count = 0
        for row in all_rows[data_start:]:
            for gi in range(num_groups):
                ci    = col_offset + gi * 3
                first = row[ci]   if ci   < len(row) else None
                last  = row[ci+1] if ci+1 < len(row) else None
                if not first or not isinstance(first, str): continue
                first = first.strip()
                if not last or not isinstance(last, str):
                    if first and 'NO MAN' not in first.upper():
                        skipped.append(f"{sheet_name}: '{first}' (no last name)")
                    continue
                last = last.strip()
                if not first or not last: continue
                if 'NO MAN' in first.upper(): continue

                all_records.append({
                    'First Name':       first,
                    'Last Name':        last,
                    'Dashboard Grade':  orig_label,
                    'Grade':            camp_grade,
                    'Gender':           'Female' if sheet_gender == 'Girl' else 'Male',
                    'Community Group':  group_label(gi),
                    'Leader':           ' & '.join(group_ldrs.get(gi, [])),
                })
                sheet_count += 1

        n_grps = len([gi for gi in group_ldrs if group_ldrs[gi]])
        parse_report.append(
            f"  ✅ {sheet_name} (Grade {grade_num} → camp {camp_grade}) "
            f"— {sheet_count} attendees, {n_grps} groups"
        )

    if skipped:
        parse_report.append(f"  ⚠ {len(skipped)} single-name entries skipped:")
        for s in skipped[:4]: parse_report.append(f"     {s}")
        if len(skipped) > 4: parse_report.append(f"     … and {len(skipped)-4} more")

    if not all_records:
        return None, parse_report + [
            "❌ No records found. Sheet names should match '8th Girls', '7th Boys', etc."
        ]

    return pd.DataFrame(all_records), parse_report

def file_ok_banner(label, df):
    st.markdown(f'<div style="padding:0 1rem;"><div class="ok-banner">✅ <strong>{label}</strong> — '
                f'{len(df)} rows · columns: {", ".join(df.columns.tolist()[:6])}'
                + (' …' if len(df.columns) > 6 else '') + '</div></div>',
                unsafe_allow_html=True)
